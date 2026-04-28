"""
File Parser — Parseia arquivos de dados enviados pelo frontend (CSV, XLSX, XLS, TXT).

Converte o conteúdo em uma lista de dicts normalizados para injeção no prompt do agente.
"""
from __future__ import annotations

import csv
import io
import logging
from typing import Any

logger = logging.getLogger(__name__)

# Mapeamento de nomes de coluna comuns → nomes canônicos
_COLUMN_ALIASES: dict[str, str] = {
    # EAN / código de barras
    "ean": "ean",
    "codigo_barras": "ean",
    "código de barras": "ean",
    "codigo de barras": "ean",
    "cod barras": "ean",
    "cod. barras": "ean",
    "gtin": "ean",
    "barcode": "ean",
    # Descrição
    "descricao": "descricao",
    "descrição": "descricao",
    "descricao do produto": "descricao",
    "descrição do produto": "descricao",
    "produto": "descricao",
    "nome": "descricao",
    "nome do produto": "descricao",
    "item": "descricao",
    # Preço unitário
    "preco_unitario": "preco_unitario",
    "preco unitario": "preco_unitario",
    "preço unitário": "preco_unitario",
    "preco": "preco_unitario",
    "preço": "preco_unitario",
    "vlr unitario": "preco_unitario",
    "vlr. unitário": "preco_unitario",
    "valor unitario": "preco_unitario",
    "valor unitário": "preco_unitario",
    "unit_price": "preco_unitario",
    # Valor unitário final/bruto → preco_unitario (planilhas com colunas separadas)
    "valor_unitario_final": "preco_unitario",
    "valor unitario final": "preco_unitario",
    "valor_unitario_bruto": "valor_unitario_bruto",
    "valor unitario bruto": "valor_unitario_bruto",
    # Quantidade
    "quantidade": "quantidade",
    "qtde": "quantidade",
    "qtd": "quantidade",
    "qtde. unitária": "quantidade",
    "qtde. unitaria": "quantidade",
    "qtde unitaria": "quantidade",
    "qtde unitária": "quantidade",
    "qty": "quantidade",
    "quantity": "quantidade",
    "quantidade_unitaria": "quantidade",
    # Valor total
    "valor_total": "valor_total",
    "valor total": "valor_total",
    "valor total do item": "valor_total",
    "vlr total": "valor_total",
    "vlr. total": "valor_total",
    "total": "valor_total",
    "valor_total_item": "valor_total",
    "valor_total_bruto": "valor_total",
    "valor total bruto": "valor_total",
    "valor_total_liquido": "valor_total_liquido",
    "valor total liquido": "valor_total_liquido",
    # Desconto / Frete
    "valor_de_desconto": "valor_desconto",
    "valor de desconto": "valor_desconto",
    "desconto": "valor_desconto",
    "valor_do_frete": "valor_frete",
    "valor do frete": "valor_frete",
    "frete": "valor_frete",
    # Data
    "data": "data_entrada",
    "data_entrada": "data_entrada",
    "data entrada": "data_entrada",
    "data/hora da entrada": "data_entrada",
    "data hora da entrada": "data_entrada",
    "dt entrada": "data_entrada",
    "date": "data_entrada",
    # Fornecedor
    "fornecedor": "fornecedor",
    "supplier": "fornecedor",
    # ICMS ST
    "valor_icms_st": "valor_icms_st",
    "valor do icms st": "valor_icms_st",
    "icms st": "valor_icms_st",
    # Outras despesas
    "valor_outras_despesas": "valor_outras_despesas",
    "valor de outras despesas": "valor_outras_despesas",
    "outras despesas": "valor_outras_despesas",
}


def _normalize_column_name(col: str) -> str:
    """Normaliza nome de coluna para o alias canônico."""
    import re as _re
    import unicodedata
    cleaned = col.strip().lower()
    # Remover acentos para matching
    cleaned_no_accent = unicodedata.normalize("NFKD", cleaned)
    cleaned_no_accent = "".join(c for c in cleaned_no_accent if not unicodedata.combining(c))
    # Substituir separadores por espaço e colapsar
    cleaned_no_accent = _re.sub(r"[_./]+", " ", cleaned_no_accent).strip()
    cleaned_no_accent = _re.sub(r"\s+", " ", cleaned_no_accent)
    # Tentar match exato com versão sem acento
    if cleaned_no_accent in _COLUMN_ALIASES:
        return _COLUMN_ALIASES[cleaned_no_accent]
    # Também tentar a versão original com acentos
    cleaned_spaces = _re.sub(r"[_./]+", " ", cleaned).strip()
    cleaned_spaces = _re.sub(r"\s+", " ", cleaned_spaces)
    if cleaned_spaces in _COLUMN_ALIASES:
        return _COLUMN_ALIASES[cleaned_spaces]
    # Tentar com underscores
    cleaned_underscore = cleaned_no_accent.replace(" ", "_")
    if cleaned_underscore in _COLUMN_ALIASES:
        return _COLUMN_ALIASES[cleaned_underscore]
    return cleaned_underscore  # retorna como está


def _parse_number(value: Any) -> float | None:
    """Converte valor para float, tratando vírgulas e strings."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(" ", "")
    if not s or s.lower() in ("none", "null", "nan", "-", ""):
        return None
    # Formato brasileiro: 1.234,56 → 1234.56
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _parse_csv(file_bytes: bytes, filename: str) -> list[dict[str, Any]]:
    """Parseia CSV ou TXT (tab-separated)."""
    # Detectar encoding
    for encoding in ("utf-8-sig", "utf-8", "latin-1", "cp1252"):
        try:
            text = file_bytes.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    else:
        text = file_bytes.decode("utf-8", errors="replace")

    # Detectar delimitador
    sample = text[:2000]
    if "\t" in sample:
        delimiter = "\t"
    elif ";" in sample and sample.count(";") > sample.count(","):
        delimiter = ";"
    else:
        delimiter = ","

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    rows = []
    for row_raw in reader:
        normalized = {}
        for col, val in row_raw.items():
            if col is None:
                continue
            canonical = _normalize_column_name(col)
            normalized[canonical] = val
        rows.append(normalized)

    return rows


def _parse_xlsx(file_bytes: bytes, filename: str) -> list[dict[str, Any]]:
    """Parseia XLSX/XLS usando openpyxl."""
    import openpyxl
    from datetime import datetime as dt

    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    # Encontrar o header: procurar a primeira linha com ≥3 colunas não-vazias
    header_row = None
    header_idx = 1
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20), 1):
        vals = [cell.value for cell in row]
        non_empty = sum(1 for v in vals if v is not None and str(v).strip())
        if non_empty >= 3:
            header_row = vals
            header_idx = i
            break

    if not header_row:
        wb.close()
        return []

    # Normalizar nomes de colunas
    col_mapping: list[tuple[int, str]] = []
    for idx, col_name in enumerate(header_row):
        if col_name and str(col_name).strip():
            canonical = _normalize_column_name(str(col_name))
            col_mapping.append((idx, canonical))

    rows = []
    for row in ws.iter_rows(min_row=header_idx + 1):
        vals = [cell.value for cell in row]
        # Pular linhas totalmente vazias
        if all(v is None or str(v).strip() == "" for v in vals):
            continue

        record: dict[str, Any] = {}
        for col_idx, canonical_name in col_mapping:
            if col_idx < len(vals):
                value = vals[col_idx]
                # Converter datetime para string
                if isinstance(value, dt):
                    value = value.strftime("%Y-%m-%d")
                record[canonical_name] = value

        # Só adicionar se tem ao menos um campo relevante
        if record.get("ean") or record.get("descricao") or record.get("valor_total"):
            rows.append(record)

    wb.close()
    return rows


def parse_uploaded_file(file_bytes: bytes, filename: str) -> list[dict[str, Any]]:
    """
    Parseia um arquivo enviado pelo usuário.

    Retorna lista de dicts com colunas normalizadas.
    Suporta: .csv, .xlsx, .xls, .txt
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext in ("xlsx", "xls"):
        rows = _parse_xlsx(file_bytes, filename)
    elif ext in ("csv", "txt", "tsv"):
        rows = _parse_csv(file_bytes, filename)
    else:
        raise ValueError(f"Formato de arquivo não suportado: .{ext}. Use CSV, XLSX, XLS ou TXT.")

    logger.info(f"Arquivo {filename} parseado: {len(rows)} registros")
    return rows


# ─── Aliases para colunas de arquivos de OFERTA ──────────────────────────────

_OFFER_PRICE_ALIASES = {
    "vda", "preco", "preço", "preco_unitario", "preço unitário", "preco unitario",
    "vlr unitario", "vlr. unitário", "valor unitario", "valor unitário",
    "unit_price", "price", "valor_total", "valor total", "vlr total",
    "vlr. total", "total", "valor_total_item",
}

_OFFER_DESC_ALIASES = {
    "descricao", "descrição", "descricao_do_produto", "descrição do produto",
    "descricao do produto", "produto", "nome", "nome_do_produto",
    "nome do produto", "item",
}

_OFFER_EAN_ALIASES = {
    "ean", "codigo_barras", "código de barras", "codigo de barras",
    "cod barras", "cod. barras", "gtin", "barcode", "cod_barras",
}


def parse_offer_file(file_bytes: bytes, filename: str) -> tuple[list[dict[str, Any]], str | None]:
    """
    Parseia um arquivo de OFERTA (XLSX/CSV) e retorna lista de itens estruturados
    no formato esperado pelo analysis engine, MAIS o fornecedor detectado.

    Returns:
        (items, fornecedor) onde:
        - items: lista de dicts com descricao, preco, ean, tipo_preco, etc.
        - fornecedor: nome do fornecedor extraído do arquivo, ou None

    Detecta automaticamente colunas de preço, descrição e EAN usando aliases.
    """
    from app.services.offer_extractor import extrair_multiplicador_inteligente
    import re as _re_offer

    # Parse the file generically first
    rows = parse_uploaded_file(file_bytes, filename)
    if not rows:
        return [], None

    # Detect which canonical columns exist
    sample_keys = set(rows[0].keys()) if rows else set()
    logger.info(f"Offer file columns: {sample_keys}")

    # ── Extract fornecedor from data or filename ──────────────────────────
    fornecedor_detectado: str | None = None

    # 1. Try 'fornecedor' column if it exists
    fornecedor_col = None
    for key in sample_keys:
        if key == "fornecedor" or _normalize_column_name(key) == "fornecedor":
            fornecedor_col = key
            break

    if fornecedor_col:
        # Get the most common non-empty value
        from collections import Counter
        forn_values = [
            str(r.get(fornecedor_col, "")).strip()
            for r in rows
            if r.get(fornecedor_col) and str(r.get(fornecedor_col, "")).strip()
        ]
        if forn_values:
            fornecedor_detectado = Counter(forn_values).most_common(1)[0][0]

    # 2. Fallback: try to extract from filename (e.g., "TABELA OL- MEDQUIMICA.xlsx")
    if not fornecedor_detectado and filename:
        # Remove extension and common prefixes
        name_clean = _re_offer.sub(r"\.(xlsx|xls|csv)$", "", filename, flags=_re_offer.IGNORECASE)
        name_clean = _re_offer.sub(r"^(tabela|oferta|promo|lista)\s*[-_]?\s*", "", name_clean, flags=_re_offer.IGNORECASE).strip()
        # Remove prefixes like "OL-" (Ofertas do Laboratório)
        name_clean = _re_offer.sub(r"^[A-Z]{1,3}\s*[-_]\s*", "", name_clean).strip()
        if name_clean and len(name_clean) >= 3:
            fornecedor_detectado = name_clean.upper()

    if fornecedor_detectado:
        logger.info(f"Fornecedor detectado do arquivo de oferta: {fornecedor_detectado}")

    # ── Find columns ──────────────────────────────────────────────────────

    # Find description column
    desc_col = None
    for key in sample_keys:
        if key in _OFFER_DESC_ALIASES or _normalize_column_name(key) == "descricao":
            desc_col = key
            break
    if not desc_col and "descricao" in sample_keys:
        desc_col = "descricao"

    # Find price column — prefer 'vda' (selling price), then 'preco_unitario', then 'valor_total'
    price_col = None
    price_priority = ["vda", "preco_unitario", "valor_unitario_final", "valor_total", "total"]
    # Build a normalized → original key map for reliable matching
    norm_to_key = {k.lower().strip(): k for k in sample_keys}
    for candidate in price_priority:
        if candidate in sample_keys:
            price_col = candidate
            break
        # Also check normalized version
        if candidate in norm_to_key:
            price_col = norm_to_key[candidate]
            break
    if not price_col:
        # Try raw key matching against aliases
        for key in sample_keys:
            norm = _normalize_column_name(key)
            if norm in _OFFER_PRICE_ALIASES or key.lower().strip() in _OFFER_PRICE_ALIASES:
                price_col = key
                break

    # Find EAN column
    ean_col = None
    for key in sample_keys:
        if key in _OFFER_EAN_ALIASES or key == "ean":
            ean_col = key
            break

    logger.info(f"Offer columns detected: desc={desc_col}, price={price_col}, ean={ean_col}")

    if not desc_col:
        logger.warning("Could not find a description column in the offer file")
        return [], fornecedor_detectado

    # Extract items
    items: list[dict[str, Any]] = []
    for row in rows:
        desc = str(row.get(desc_col, "")).strip()
        if not desc or len(desc) < 3:
            continue

        # Parse price
        preco = None
        if price_col:
            raw_price = row.get(price_col)
            preco = _parse_number(raw_price)
            if preco is not None and preco <= 0:
                preco = None

        # Parse EAN
        ean = None
        if ean_col:
            raw_ean = row.get(ean_col)
            if raw_ean is not None:
                ean_str = str(raw_ean).strip()
                # Remove .0 from float-parsed EANs
                if ean_str.endswith(".0"):
                    ean_str = ean_str[:-2]
                # Validate: EAN should be numeric and 8-14 digits
                if ean_str.isdigit() and 8 <= len(ean_str) <= 14:
                    ean = ean_str

        # Determine multiplicador
        mult = extrair_multiplicador_inteligente(desc)

        items.append({
            "descricao": desc,
            "preco": preco,
            "ean": ean,
            "tipo_preco": "absoluto" if preco else "sem_preco",
            "desconto_percentual": None,
            "multiplicador_embalagem": mult,
        })

    logger.info(f"Offer file parsed: {len(items)} items extracted (desc={desc_col}, price={price_col}, ean={ean_col})")
    return items, fornecedor_detectado


def _calcular_preco_unitario(row: dict) -> float | None:
    """
    Calcula preço unitário priorizando campos mais específicos:
    1. preco_unitario (valor_unitario_final mapeado aqui via aliases)
    2. valor_unitario_bruto (se final não existir)
    3. valor_total_liquido / quantidade
    4. valor_total / quantidade
    5. valor_total sozinho (qtde=1 implícita)
    """
    # 1. Preço unitário direto (inclui valor_unitario_final via alias)
    pu = _parse_number(row.get("preco_unitario"))
    if pu and pu > 0:
        return round(pu, 4)

    # 2. Valor unitário bruto (antes de descontos)
    pu_bruto = _parse_number(row.get("valor_unitario_bruto"))
    if pu_bruto and pu_bruto > 0:
        return round(pu_bruto, 4)

    qtde = _parse_number(row.get("quantidade"))

    # 3. Total líquido / quantidade (mais preciso que total bruto)
    vt_liq = _parse_number(row.get("valor_total_liquido"))
    if vt_liq and qtde and qtde > 0:
        return round(vt_liq / qtde, 4)

    # 4. Valor total / quantidade
    valor_total = _parse_number(row.get("valor_total"))
    if valor_total and qtde and qtde > 0:
        return round(valor_total / qtde, 4)

    # 5. Fallback: valor_total é o preço unitário (qtde=1 implícita)
    if valor_total and valor_total > 0:
        return round(valor_total, 4)

    return None


def format_file_data_for_prompt(rows: list[dict[str, Any]], max_rows: int = 300, texto_oferta: str | None = None) -> str:
    """
    Formata os dados parseados como texto estruturado para injeção no prompt do agente.

    Se texto_oferta for fornecido, faz pré-matching inteligente: cruza os itens da oferta
    com os dados do arquivo usando fuzzy matching e retorna apenas os dados relevantes
    já com estatísticas calculadas (menor preço, média, demanda, etc.).

    Sem texto_oferta, retorna um resumo geral dos dados agrupados por EAN.
    """
    from collections import defaultdict
    from statistics import mean
    import unicodedata
    import re as _re

    if not rows:
        return "Nenhum dado encontrado no arquivo enviado."

    # ─── Agrupar por EAN e calcular estatísticas ──────────────────────────────
    by_ean: dict[str, list[dict]] = defaultdict(list)
    sem_ean: list[dict] = []

    for row in rows:
        ean = row.get("ean")
        if ean and str(ean).strip() and str(ean).strip() not in ("0", "None", "none", "null"):
            by_ean[str(ean).strip()].append(row)
        elif row.get("descricao"):
            sem_ean.append(row)

    # ─── Calcular estatísticas por EAN ────────────────────────────────────────
    def _stats_for_entries(entries: list[dict]) -> dict:
        precos = []
        datas = []
        descricao = ""
        qtde_total = 0.0

        for e in entries:
            if not descricao and e.get("descricao"):
                descricao = str(e["descricao"]).strip()

        # Normalizar para preco por "pílula/ml" extraindo o multiplicador
        from app.services.offer_extractor import extrair_multiplicador_inteligente
        mult = extrair_multiplicador_inteligente(descricao) if descricao else 1.0
        if mult <= 0:
            mult = 1.0

        for e in entries:
            pu = _calcular_preco_unitario(e)
            if pu and pu > 0:
                precos.append(round(pu / mult, 4))
            d = e.get("data_entrada")
            if d:
                datas.append(str(d))
            qtde_total += _parse_number(e.get("quantidade")) or 0

        menor = min(precos) if precos else None
        media = round(mean(precos), 4) if precos else None
        maior = max(precos) if precos else None
        ultima_data = max(datas) if datas else "N/A"
        primeira_data = min(datas) if datas else "N/A"

        return {
            "descricao": descricao,
            "qtd_entradas": len(entries),
            "qtde_total": qtde_total,
            "menor_preco": menor,
            "media_preco": media,
            "maior_preco": maior,
            "primeira_data": primeira_data,
            "ultima_data": ultima_data,
        }

    ean_stats: dict[str, dict] = {}
    for ean, entries in by_ean.items():
        ean_stats[ean] = _stats_for_entries(entries)
        ean_stats[ean]["ean"] = ean

    # ─── Se não tem texto da oferta, retorna resumo geral (truncado) ──────────
    if not texto_oferta:
        return _format_general_summary(ean_stats, len(rows), max_rows, sem_ean)

    # ─── PRÉ-MATCHING: cruzar itens da oferta com dados do arquivo ────────────
    matched_items = _match_offer_items_to_file(texto_oferta, ean_stats)
    return _format_matched_data(matched_items, len(rows), len(ean_stats))


def _slugify(text: str) -> str:
    """Remove acentos, converte para minúsculas, remove caracteres especiais."""
    import unicodedata
    import re as _re
    text = unicodedata.normalize("NFKD", text.lower())
    text = "".join(c for c in text if not unicodedata.combining(c))
    text = _re.sub(r"[^a-z0-9\s]", " ", text)
    return _re.sub(r"\s+", " ", text).strip()


def _extract_tokens(text: str) -> set[str]:
    """Extrai tokens significativos (>= 2 chars) de uma string."""
    return {t for t in _slugify(text).split() if len(t) >= 2}


def _match_offer_items_to_file(
    texto_oferta: str,
    ean_stats: dict[str, dict],
) -> list[dict]:
    """
    Cruza os itens de uma oferta com os dados do arquivo usando fuzzy matching.

    Retorna lista de dicts com dados do match para cada item da oferta encontrado.
    """
    import re as _re

    # Construir índice de busca: tokens -> lista de EANs
    token_index: dict[str, list[str]] = {}
    for ean, stats in ean_stats.items():
        desc = stats.get("descricao", "")
        tokens = _extract_tokens(desc)
        for token in tokens:
            if token not in token_index:
                token_index[token] = []
            token_index[token].append(ean)

    # Extrair possíveis itens da oferta (cada linha não-vazia pode ser um item)
    linhas = [l.strip() for l in texto_oferta.split("\n") if l.strip()]

    # Filtrar linhas que parecem ser itens (contêm descrição + preço)
    offer_items = []
    for linha in linhas:
        # Linha com preço (R$, = XX, XX,XX etc.)
        has_price = bool(_re.search(r"(?:R\$\s*)?(\d+[.,]\d{2})", linha))
        has_desc = len(linha) > 5 and any(c.isalpha() for c in linha)
        if has_desc and has_price:
            offer_items.append(linha)

    if not offer_items:
        # Fallback: usar todas as linhas com alguma substância
        offer_items = [l for l in linhas if len(l) > 8 and any(c.isalpha() for c in l)]

    matched_results = []
    used_eans = set()  # Evitar que o mesmo EAN seja matcheado para múltiplos itens

    for item_text in offer_items:
        # Extrair tokens — incluir tokens de 2+ chars para termos farmacêuticos (mg, ml)
        item_tokens = {t for t in _slugify(item_text).split() if len(t) >= 2}
        if len(item_tokens) < 1:
            continue

        # Separar tokens por tipo para scoring ponderado
        # Tokens longos/alfabéticos (nomes de fármacos) têm peso maior
        drug_tokens = {t for t in item_tokens if len(t) >= 4 and t.isalpha()}
        generic_tokens = item_tokens - drug_tokens  # mg, ml, números, etc.

        # Buscar EANs candidatos com scoring ponderado
        candidate_scores: dict[str, float] = {}
        for token in drug_tokens:
            for ean in token_index.get(token, []):
                if ean not in used_eans:
                    candidate_scores[ean] = candidate_scores.get(ean, 0) + 5.0  # peso alto
        for token in generic_tokens:
            for ean in token_index.get(token, []):
                if ean not in used_eans:
                    candidate_scores[ean] = candidate_scores.get(ean, 0) + 1.0  # peso baixo

        if not candidate_scores:
            # Nenhum match encontrado
            matched_results.append({
                "item_oferta": item_text,
                "matched": False,
            })
            continue

        # Ordenar por score ponderado (tokens de fármaco valem mais)
        sorted_candidates = sorted(candidate_scores.items(), key=lambda x: x[1], reverse=True)
        best_ean, best_score = sorted_candidates[0]

        # Verificar se o melhor match tem ao menos 1 token de fármaco em comum
        best_desc_tokens = _extract_tokens(ean_stats[best_ean].get("descricao", ""))
        common_drug_tokens = drug_tokens & best_desc_tokens
        
        # Exigir ao menos 1 token de fármaco (4+ chars, alfabético) em comum
        # OU ao menos score >= 10 (vários tokens genéricos em comum)
        if not common_drug_tokens and best_score < 10:
            matched_results.append({
                "item_oferta": item_text,
                "matched": False,
            })
            continue

        stats = ean_stats[best_ean]
        n_common_drug = len(common_drug_tokens)
        confianca = "alto" if n_common_drug >= 2 else "medio" if n_common_drug >= 1 else "baixo"

        used_eans.add(best_ean)  # Marcar EAN como usado

        # Calcular demanda mensal estimada
        demanda_mes = 0.0
        if stats["primeira_data"] != "N/A" and stats["ultima_data"] != "N/A":
            try:
                from datetime import datetime
                d1 = datetime.strptime(str(stats["primeira_data"])[:10], "%Y-%m-%d")
                d2 = datetime.strptime(str(stats["ultima_data"])[:10], "%Y-%m-%d")
                meses = max(1, (d2 - d1).days / 30)
                demanda_mes = round(stats["qtde_total"] / meses, 1)
            except Exception:
                pass

        matched_results.append({
            "item_oferta": item_text,
            "matched": True,
            "ean": best_ean,
            "descricao_arquivo": stats["descricao"],
            "confianca_match": confianca,
            "tokens_match": best_score,
            "menor_preco_unitario": stats["menor_preco"],
            "media_preco_unitario": stats["media_preco"],
            "maior_preco_unitario": stats["maior_preco"],
            "qtd_entradas": stats["qtd_entradas"],
            "qtde_total": stats["qtde_total"],
            "demanda_mes": demanda_mes,
            "primeira_data": stats["primeira_data"],
            "ultima_data": stats["ultima_data"],
        })

        logger.debug(
            f"Match oferta->arquivo: '{item_text[:40]}' -> EAN {best_ean} "
            f"({stats['descricao'][:40]}) score={best_score} confianca={confianca}"
        )

    return matched_results


def _format_matched_data(
    matched_items: list[dict],
    total_registros: int,
    total_eans: int,
) -> str:
    """Formata os resultados do pré-matching como texto estruturado para o prompt."""
    lines = [
        f"=== DADOS DO ARQUIVO CRUZADOS COM A OFERTA ===",
        f"(Arquivo original: {total_registros} registros, {total_eans} EANs únicos)",
        "",
    ]

    matched_count = sum(1 for m in matched_items if m.get("matched"))
    lines.append(f"Itens da oferta com match no arquivo: {matched_count}/{len(matched_items)}")
    lines.append("")

    for idx, item in enumerate(matched_items, 1):
        lines.append(f"--- ITEM {idx} DA OFERTA ---")
        lines.append(f"Texto original: {item['item_oferta']}")

        if item.get("matched"):
            s = item
            menor_str = f"R${s['menor_preco_unitario']:.2f}" if s.get('menor_preco_unitario') else "N/A"
            media_str = f"R${s['media_preco_unitario']:.2f}" if s.get('media_preco_unitario') else "N/A"
            maior_str = f"R${s['maior_preco_unitario']:.2f}" if s.get('maior_preco_unitario') else "N/A"

            lines.append(f"MATCH ENCONTRADO (confiança: {s['confianca_match']})")
            lines.append(f"  EAN: {s['ean']}")
            lines.append(f"  Descrição no arquivo: {s['descricao_arquivo']}")
            lines.append(f"  Histórico de preços unitários:")
            lines.append(f"    Menor preço pago: {menor_str}")
            lines.append(f"    Média de preço pago: {media_str}")
            lines.append(f"    Maior preço pago: {maior_str}")
            lines.append(f"  Entradas: {s['qtd_entradas']} compras, {s['qtde_total']:.0f} unidades total")
            lines.append(f"  Demanda mensal estimada: {s['demanda_mes']:.1f} un/mês")
            lines.append(f"  Período: {s['primeira_data']} até {s['ultima_data']}")
        else:
            lines.append(f"SEM MATCH no arquivo - produto não encontrado no histórico")

        lines.append("")

    lines.append("=== FIM DOS DADOS CRUZADOS ===")
    return "\n".join(lines)


def _format_general_summary(
    ean_stats: dict[str, dict],
    total_rows: int,
    max_rows: int,
    sem_ean: list[dict],
) -> str:
    """Formata resumo geral (sem oferta para cruzar) — fallback."""
    from statistics import mean

    lines = [
        f"=== DADOS DO ARQUIVO DO CLIENTE ({total_rows} registros, {len(ean_stats)} EANs únicos) ===",
        "",
        "Formato: EAN | Descrição | Qtd Entradas | Menor Preço Unit. | Média Preço Unit. | Última Data",
        "---",
    ]

    count = 0
    for ean in sorted(ean_stats.keys()):
        if count >= max_rows:
            lines.append(f"... e mais {len(ean_stats) - count} EANs não exibidos")
            break

        s = ean_stats[ean]
        menor_str = f"R${s['menor_preco']:.2f}" if s.get('menor_preco') else "N/A"
        media_str = f"R${s['media_preco']:.2f}" if s.get('media_preco') else "N/A"

        lines.append(
            f"EAN: {ean} | {s['descricao'][:60]} | {s['qtd_entradas']} entradas ({s['qtde_total']:.0f} un) | "
            f"Menor: {menor_str} | Média: {media_str} | Última: {s['ultima_data']}"
        )
        count += 1

    if sem_ean:
        lines.append(f"\n--- {len(sem_ean)} registros sem EAN (ignorados para comparação) ---")

    lines.append("\n=== FIM DOS DADOS DO ARQUIVO ===")
    return "\n".join(lines)

